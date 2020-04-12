from openpyxl import load_workbook, Workbook
from datetime import date, timedelta

workbook = load_workbook(filename="sheet_name.xlsx")
# workbook = Workbook()
sheet = workbook.active

sdate = date(2020, 4, 1)   # start date
edate = date(2020, 4, 30)   # end date

delta = edate - sdate       # as timedelta
k=2
for i in range(delta.days + 1):
	day = sdate + timedelta(days=i)
	if day.strftime("%A") != 'Sunday' and day.strftime("%A") !='Saturday':
		sheet.cell(row=k, column=1).value = day
		k=k+2

#print(sheet["A:B"])
# for value in sheet.iter_rows(min_row=1, min_col=1, max_col=1, values_only=True):
# 	print(value)

# for i in range(1,10):
# 	sheet.cell(row=i, column=1).value = 2

	# print(day)
# print(type(date))

workbook.save(filename="sheet_name.xlsx")

